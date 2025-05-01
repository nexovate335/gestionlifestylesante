from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from .models import CustomUser, Personnel
import csv
from django.http import HttpResponse
from django.utils.text import slugify
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO
from datetime import datetime
from django.conf import settings

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


# Export Excel
def export_as_excel(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['deleted_at', 'id']]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = slugify(meta.verbose_name_plural)

    # En-têtes en gras
    for col_num, field_name in enumerate(field_names, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = field_name
        cell.font = Font(bold=True)

    # Données
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(field_names, 1):
            value = getattr(obj, field_name)
            if field_name == 'date' and value:
                value = value.strftime('%Y-%m-%d %H:%M')
            worksheet.cell(row=row_num, column=col_num, value=str(value))

    # Largeur colonnes
    for col_num, field_name in enumerate(field_names, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = max(20, len(field_name) + 5)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{slugify(meta.verbose_name_plural)}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

export_as_excel.short_description = "Exporter en Excel"


# Export CSV
def export_as_csv(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['deleted_at', 'id']]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={slugify(meta.verbose_name_plural)}.csv'
    
    writer = csv.writer(response)
    writer.writerow(field_names)

    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if field == 'date' and value:
                value = value.strftime('%Y-%m-%d %H:%M')
            row.append(str(value))
        writer.writerow(row)

    return response

export_as_csv.short_description = "Exporter en CSV"


def export_as_pdf(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['deleted_at', 'id']]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_lifestyle.jpg'
    logo = Image(logo_path, width=100, height=50)
    elements.append(logo)

    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Export de {meta.verbose_name_plural.title()}</b>", styles['Title'])
    date_gen = Paragraph(f"Date de génération : {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    elements.extend([title, date_gen, Spacer(1, 12)])

    data = [field_names]

    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M')
            row.append(str(value))
        data.append(row)

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ]))
    elements.append(table)

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Exporté automatiquement depuis l'administration Lifestyle.", styles['Italic']))

    doc.build(elements)
    buffer.seek(0)

    filename = f"{slugify(meta.verbose_name_plural)}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
export_as_pdf.short_description = "Exporter en PDF"


# -------------------- ADMIN --------------------

class BaseAdmin(admin.ModelAdmin):
    actions = [export_as_csv, export_as_excel, export_as_pdf]
    list_filter = ('deleted_at',)

# Inline de Personnel dans CustomUser
class PersonnelInline(admin.StackedInline):
    model = Personnel
    can_delete = False
    verbose_name_plural = 'Personnel'


# Admin personnalisé pour CustomUser
@admin.register(CustomUser)
class CustomUserAdmin(BaseAdmin):
    inlines = (PersonnelInline,)
    
    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        ('Informations personnelles', {'fields': ('last_name', 'first_name', 'email')}),
        ('Permissions', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Dates importantes', {'fields': ('last_login', 'date_joined')}),
    )
    list_display = ('username', 'email', 'last_name', 'first_name', 'is_staff')
    search_fields = ('username', 'email', 'first_name', 'last_name')
    list_filter = ('is_staff', 'is_superuser', 'is_active', 'groups')


# Admin pour Personnel
@admin.register(Personnel)
class PersonnelAdmin(BaseAdmin):
    list_display = ('user', 'last_name', 'first_name', 'email', 'sexe', 'fonction', 'telephone', 'deleted_at')
    search_fields = ('last_name', 'first_name', 'user__last_name', 'user__first_name', 'user__email', 'telephone', 'fonction')
    list_filter = ('sexe', 'fonction', 'deleted_at')
    
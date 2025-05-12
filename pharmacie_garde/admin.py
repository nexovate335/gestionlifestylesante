from django.contrib import admin
from .models import PhGardeProduit, PhGardeCommande, PhGardeStock, PhGardeVente, PhGardeFacturePharmacie, PhGardeFactureAvance
import csv
from django.http import HttpResponse
from django.utils.text import slugify
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO
from datetime import datetime
from django.conf import settings

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


# Export Excel
def export_as_excel(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['deleted_at', 'id']]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = slugify(meta.verbose_name_plural)

    # En-têtes en gras
    for col_num, field_name in enumerate(field_names, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = field_name
        cell.font = Font(bold=True)

    # Données
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(field_names, 1):
            value = getattr(obj, field_name)
            if field_name == 'date' and value:
                value = value.strftime('%Y-%m-%d %H:%M')
            worksheet.cell(row=row_num, column=col_num, value=str(value))

    # Largeur colonnes
    for col_num, field_name in enumerate(field_names, 1):
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = max(20, len(field_name) + 5)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{slugify(meta.verbose_name_plural)}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

export_as_excel.short_description = "Exporter en Excel"


# Export CSV
def export_as_csv(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['deleted_at', 'id']]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename={slugify(meta.verbose_name_plural)}.csv'
    
    writer = csv.writer(response)
    writer.writerow(field_names)

    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if field == 'date' and value:
                value = value.strftime('%Y-%m-%d %H:%M')
            row.append(str(value))
        writer.writerow(row)

    return response

export_as_csv.short_description = "Exporter en CSV"


# Export PDF
def export_as_pdf(modeladmin, request, queryset):
    meta = modeladmin.model._meta
    field_names = [field.name for field in meta.fields if field.name not in ['deleted_at', 'id']]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Ajouter un logo en haut à gauche
    logo_path = settings.BASE_DIR / 'static' / 'images' / 'logo_lifestyle.jpg'  # Assure-toi que le chemin d'accès au logo est correct
    logo = Image(logo_path, width=100, height=50)  # Ajuste la taille du logo
    elements.append(logo)

    # Titre et date de génération
    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Export de {meta.verbose_name_plural.title()}</b>", styles['Title'])
    date_gen = Paragraph(f"Date de génération : {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
    elements.extend([title, date_gen, Spacer(1, 12)])

    # Données à exporter
    data = [field_names]
    total_montant = 0

    for obj in queryset:
        row = []
        for field in field_names:
            value = getattr(obj, field)
            if field == 'date' and value:
                value = value.strftime('%Y-%m-%d %H:%M')
            if field == "montant":
                try:
                    total_montant += float(value)
                except (ValueError, TypeError):
                    pass
            row.append(str(value))
        data.append(row)

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Total général des montants :</b> {total_montant:.2f} FCFA", styles['Normal']))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Exporté automatiquement depuis l'administration Lifestyle.", styles['Italic']))

    doc.build(elements)
    buffer.seek(0)

    filename = f"{slugify(meta.verbose_name_plural)}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response

export_as_pdf.short_description = "Exporter en PDF"

# Base Admin pour les actions CSV, Excel, PDF
class BaseAdmin(admin.ModelAdmin):
    actions = [export_as_csv, export_as_excel, export_as_pdf]


@admin.register(PhGardeProduit)
class PhGardeProduitAdmin(BaseAdmin):
    list_display = ('nom_produit',)
    search_fields = ('nom_produit',)
    ordering = ('nom_produit',)

# Admin pour Commande
@admin.register(PhGardeCommande)
class PhGardeCommandeAdmin(BaseAdmin):
    list_display = (
        'produit', 
        'quantite_commande', 
        'prix', 
        'total_achat', 
        'fournisseur', 
        'date_admission', 
        'date_expiration', 
        'date_commande'
    )
    list_filter = ('produit', 'fournisseur', 'date_commande', 'date_admission', 'date_expiration')
    search_fields = ('produit__nom_produit', 'fournisseur__nom_fournisseur')
    ordering = ('date_commande',)



# Admin pour Stock
@admin.register(PhGardeStock)
class PhGardeStockAdmin(BaseAdmin):
    list_display = ('produit', 'quantite_reelle', 'prix_unitaire', 'prix_total', 'date_stock')
    search_fields = ('produit__nom_produit',)
    ordering = ('produit__nom_produit',)


# Admin pour Vente
@admin.register(PhGardeVente)
class PhGardeVenteAdmin(BaseAdmin):
    list_display = ('produit', 'quantite_vendue', 'prix_unitaire', 'prix_total', 'date_vente')
    list_filter = ('produit', 'date_vente')
    search_fields = ('produit__nom_produit',)
    ordering = ('date_vente',)


# Admin pour FacturePharmacie
@admin.register(PhGardeFacturePharmacie)
class PhGardeFacturePharmacieAdmin(BaseAdmin):
    list_display = ('patient','nom_personne', 'facture_date_time', 'total', 'total_somme_verse', 'avance', 'reste_a_payer', 'nombre_versements', 'paye', 'deleted_at')
    list_filter = ('paye', 'facture_date_time', 'deleted_at')
    search_fields = ('patient__nom', 'save_by__nom')
    readonly_fields = ('total_somme_verse', 'avance', 'reste_a_payer', 'paye', 'nombre_versements')


# Admin pour FactureAvance
@admin.register(PhGardeFactureAvance)
class PhGardeFactureAvanceAdmin(BaseAdmin):
    list_display = ('facture', 'montant', 'date_versement', 'deleted_at')
    list_filter = ('date_versement', 'deleted_at')
    search_fields = ('facture__patient__nom', 'facture__patient__prenom', 'facture__id')
    readonly_fields = ('montant', 'date_versement')  # Champs en lecture seule

    